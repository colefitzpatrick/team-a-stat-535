# -*- coding: utf-8 -*-
"""
Created on Sun Nov  8 09:10:06 2020

@author: Cole Fitzpatrick
"""


import pandas as pd
import openpyxl

wb = openpyxl.load_workbook('variablegroups.xlsx')
ws = wb["Sheet1"]

df = pd.read_excel("owid-covid-vif.xlsx", sheet_name = "Sheet2")
col_list = []

"""put the column names in a list"""
for col in df.columns:
    col_list.append(col)
    
col_list.remove('location')

"""create correlation matrix"""
corr_df = df.corr()

"""iterate through the columns, create a df that contains the correlated values (score > 0.40), writes the variables that are correlated to column in an excel file associated with the column"""
column_num = 1
for data in col_list:
    correlated = []
    row_num = 2
    correlated_df = corr_df[(corr_df[data] > 0.40) & (corr_df[data] != 1)][data]
    correlated = list(correlated_df.index)
    ws.cell(row=1,column=column_num).value = data
    for row in correlated:
        ws.cell(row=row_num,column=column_num).value = row
        row_num += 1
    column_num += 1
    
wb.save('variablegroups.xlsx')
    
    


